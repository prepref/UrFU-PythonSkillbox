import pandas as pd
import openpyxl as op
import matplotlib.pyplot as plt

def to_correct_pd(file_name:str) -> (pd.DataFrame):

    new_df = pd.read_csv(file_name,names=["name","salary_from","salary_to","salary_currency","area_name","published_at"])
    new_df['salary'] = new_df[["salary_from", "salary_to"]].sum(axis=1).fillna(0)
    new_df.loc[(((new_df['salary'] != new_df['salary_to']) & (new_df['salary'] == new_df['salary_from'])) | ((new_df['salary'] != new_df['salary_from']) & (new_df['salary'] == new_df['salary_to'])))==0,'salary'] = (new_df['salary_to']+new_df['salary_from'])/2
    new_df['published_at'] = new_df['published_at'].str.extract(r'(\d{4})').astype(int)

    return new_df

def to_correct_pd_profession(new_df:pd.DataFrame,vac_name:str) -> (pd.DataFrame):
    return new_df[new_df['name'].str.lower().str.contains(vac_name.lower(), na=False)]

def salary_by_year(new_df:pd.DataFrame) -> dict:
    res = dict()
    res = new_df.groupby(['published_at']).agg({'salary':'mean'}).astype(float).sort_values(by=['published_at'],ascending=True).to_dict()['salary']
    return {key:round(value) for key,value in res.items()}

def number_of_vacancies_by_year(new_df:pd.DataFrame) -> dict:
    res = dict()
    res = new_df.groupby(['published_at']).agg({'name':'count'}).sort_values(by=['published_at'],ascending=True).to_dict()['name']
    return res

def salary_by_year_profession(new_df: pd.DataFrame, profession_df:pd.DataFrame) -> dict:
    res = dict()
    profession_dict = profession_df.groupby(['published_at']).agg({'salary':'mean'}).astype(float).sort_values(by=['published_at'],ascending=True).to_dict()['salary']
    res = new_df.groupby(['published_at']).agg({'salary':'mean'}).astype(int).sort_values(by=['published_at'],ascending=True).to_dict()['salary']
    return {key:round(profession_dict[key]) if key in profession_dict.keys() else 0 for key in res }

def number_of_vacancies_by_year_profession(new_df: pd.DataFrame, profession_df:pd.DataFrame) -> dict:
    res = dict()
    profession_dict = profession_df.groupby(['published_at']).agg({'name':'count'}).astype(int).sort_values(by=['published_at'],ascending=True).to_dict()['name']
    res = new_df.groupby(['published_at']).agg({'name':'count'}).sort_values(by=['published_at'],ascending=True).to_dict()['name']
    return {key:profession_dict[key] if key in profession_dict.keys() else 0 for key in res }   

def salary_by_area(new_df:pd.DataFrame) -> dict:
    res = dict()
    df_area_salaries = new_df.groupby(['area_name']).agg({'salary':'mean', 'name':'count'}).astype(float).sort_values(by=['salary','area_name'],ascending=(False,True))
    df_area_salaries['name'] = df_area_salaries['name']/df_area_salaries['name'].sum()
    df_area_salaries = df_area_salaries[df_area_salaries.name*100 >= 1]             
    res = {key:round(value) for key,value in list(df_area_salaries.to_dict()['salary'].items())[:10]}
    return res

def  number_of_vacancies_by_area(new_df:pd.DataFrame) -> dict:
    res = dict()
    df_area_vacancies = new_df.groupby(['area_name']).agg({'name':'count'}).sort_values(by=['name','area_name'],ascending=(False,True))
    df_area_vacancies['name'] = df_area_vacancies['name']/df_area_vacancies['name'].sum() * 100
    df_area_vacancies = df_area_vacancies[df_area_vacancies.name*100 >= 1]   
    res = {key:round(value,2) for key,value in list(df_area_vacancies.to_dict()['name'].items())[:10]}
    return res

def create_report() -> None:
    file_name = 'vacancies.csv'
    new_df = to_correct_pd(file_name)

    salary_year = salary_by_year(new_df)
    vacansies_year = number_of_vacancies_by_year(new_df)
    salary_area = salary_by_area(new_df)
    vacansies_area = number_of_vacancies_by_area(new_df)

    list_tab1 = [[key, round(salary_year[key]), vacansies_year[key]] for key in salary_year]
    list_tab2 = [[key, round(salary_area[key])] for key in salary_area]
    list_tab3 = [[key, vacansies_area[key]] for key in vacansies_area]
    list_tab1.insert(0,["Год","Средняя зарплата","Количество вакансий"])
    list_tab2.insert(0,["Город","Уровень зарплат","Доля вакансий, %"])
    list_tab3.insert(0,["Город","Доля вакансий, %"])

    excel_doc = op.Workbook()
    excel_doc.create_sheet(title='Статистика по годам', index=0)
    excel_doc.create_sheet(title='Статистика по городам', index=1)

    sheetnames = excel_doc.sheetnames
    excel_doc.remove(excel_doc[sheetnames[2]])

    i=1
    for row in list_tab1:
        if i == 1:
            excel_doc[sheetnames[0]][f"A{i}"] = str(row[0])
            excel_doc[sheetnames[0]][f"B{i}"] = str(row[1])
            excel_doc[sheetnames[0]][f"C{i}"] = str(row[2])
        else:
            excel_doc[sheetnames[0]][f"A{i}"] = row[0]
            excel_doc[sheetnames[0]][f"B{i}"] = row[1]
            excel_doc[sheetnames[0]][f"C{i}"] = row[2]
        i+=1

    i=1
    for row in list_tab2:
        if i == 1:
            excel_doc[sheetnames[1]][f"A{i}"] = str(row[0])
            excel_doc[sheetnames[1]][f"B{i}"] = str(row[1])
        else:
            excel_doc[sheetnames[1]][f"A{i}"] = row[0]
            excel_doc[sheetnames[1]][f"B{i}"] = row[1]
        i+=1
    
    i=1
    for row in list_tab3:
        if i == 1:
            excel_doc[sheetnames[1]][f"D{i}"] = str(row[0])
            excel_doc[sheetnames[1]][f"E{i}"] = str(row[1])
        else:
            excel_doc[sheetnames[1]][f"D{i}"] = row[0]
            excel_doc[sheetnames[1]][f"E{i}"] = row[1]
        i+=1


    excel_doc.save('student_works/report.xlsx')

def create_plot():
    file_name = 'vacancies.csv'
    vac_name = input()

    new_df= to_correct_pd(file_name)
    profession_df = to_correct_pd_profession(new_df,vac_name)

    salary_year = salary_by_year(new_df)
    vacansies_year = number_of_vacancies_by_year(new_df)
    salary_year_profession = salary_by_year_profession(new_df,profession_df)
    vacansies_year_profession = number_of_vacancies_by_year_profession(new_df,profession_df)
    salary_area = salary_by_area(new_df)
    vacansies_area = number_of_vacancies_by_area(new_df)

    fig, sub = plt.subplots(2, 2)

    sub[0,0].bar(salary_year.keys(),salary_year.values(),label='средняя з/п')
    sub[0,0].bar(salary_year.keys(),salary_year_profession.values(),label=f'з/п {vac_name}')
    sub[0,0].set_title('Уровень зарплат по годам',fontsize=8)
    sub[0,0].grid(axis='y')
    sub[0,0].set_xticks(ticks=list(salary_year.keys()),labels=list(salary_year.keys()),rotation=90,fontsize=8)
    sub[0,0].legend(['средняя з/п',f'з/п {vac_name}'],fontsize=8)
    sub[0,0].tick_params(axis='y', labelsize=8)

    sub[0,1].bar(vacansies_year.keys(),vacansies_year.values(),label='Количество вакансий')
    sub[0,1].bar(vacansies_year.keys(),vacansies_year_profession.values(),label=f'Количество вакансий {vac_name}')
    sub[0,1].set_title('Количество вакансий по годам',fontsize=8)
    sub[0,1].grid(axis='y')
    sub[0,1].set_xticks(ticks=list(vacansies_year.keys()),labels=list(vacansies_year.keys()),rotation=90,fontsize=8)
    sub[0,1].legend(['Количество вакансий',f'Количество вакансий {vac_name}'],fontsize=8)
    sub[0,1].tick_params(axis='y', labelsize=8)

    l=list()
    for i in list(reversed(salary_area.keys())):
        if '-' in i:
            l.append(i.replace('-','-\n'))
        elif ' ' in i:
            l.append(i.replace(' ','\n'))
        else:
            l.append(i)
    sub[1,0].barh(list(salary_area.keys()),list(salary_area.values()))
    sub[1,0].invert_yaxis()
    sub[1,0].set_title('Уровень зарплат по городам',fontsize=8)
    sub[1,0].tick_params(axis='y', labelsize=6)
    sub[1,0].tick_params(axis='x', labelsize=8)
    sub[1,0].grid(axis='x')
    sub[1,0].set_yticklabels(l, fontdict={'horizontalalignment': 'right',
                                            'verticalalignment': 'center'})

    vacansies_area["Другие"] = 100-sum(vacansies_area.values())
    sub[1,1].pie(vacansies_area.values(),labels=vacansies_area.keys(),textprops={'fontsize': 6})
    sub[1,1].set_title('Доля вакансий по городам',fontsize=8)
 
    plt.show()

    return sub

if __name__=="__main__":
    create_report()

 