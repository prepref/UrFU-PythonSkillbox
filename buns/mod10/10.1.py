import pandas as pd
import openpyxl as op

def to_correct_pd(file_name:str) -> (pd.DataFrame):

    new_df = pd.read_csv(file_name,names=["name","salary_from","salary_to","salary_currency","area_name","published_at"])
    new_df['salary'] = new_df[["salary_from", "salary_to"]].sum(axis=1).fillna(0)
    new_df.loc[(((new_df['salary'] != new_df['salary_to']) & (new_df['salary'] == new_df['salary_from'])) | ((new_df['salary'] != new_df['salary_from']) & (new_df['salary'] == new_df['salary_to'])))==0,'salary'] = (new_df['salary_to']+new_df['salary_from'])/2
    new_df['published_at'] = new_df['published_at'].str.extract(r'(\d{4})').astype(int)

    return new_df

def salary_by_year(new_df:pd.DataFrame) -> dict:
    res = dict()
    res = new_df.groupby(['published_at']).agg({'salary':'mean'}).astype(float).sort_values(by=['published_at'],ascending=True).to_dict()['salary']
    return res

def number_of_vacancies_by_year(new_df:pd.DataFrame) -> dict:
    res = dict()
    res = new_df.groupby(['published_at']).agg({'name':'count'}).sort_values(by=['published_at'],ascending=True).to_dict()['name']
    return res

def salary_by_area(new_df:pd.DataFrame) -> dict:
    res = dict()
    df_area_salaries = new_df.groupby(['area_name']).agg({'salary':'mean', 'name':'count'}).astype(float).sort_values(by=['salary','area_name'],ascending=(False,True))
    df_area_salaries['name'] = df_area_salaries['name']/df_area_salaries['name'].sum()
    df_area_salaries = df_area_salaries[df_area_salaries.name*100 >= 1]             
    res = {key:value for key,value in list(df_area_salaries.to_dict()['salary'].items())[:10]}
    return res

def  number_of_vacancies_by_area(new_df:pd.DataFrame) -> dict:
    res = dict()
    df_area_vacancies = new_df.groupby(['area_name']).agg({'name':'count'}).sort_values(by=['name','area_name'],ascending=(False,True))
    df_area_vacancies['name'] = df_area_vacancies['name']/df_area_vacancies['name'].sum() * 100
    df_area_vacancies = df_area_vacancies[df_area_vacancies.name*100 >= 1]   
    res = {key:round(value,2) for key,value in list(df_area_vacancies.to_dict()['name'].items())[:10]}
    return res

def create_report() -> None:
    file_name = 'vacancies.csv'
    new_df = to_correct_pd(file_name)

    salary_year = salary_by_year(new_df)
    vacansies_year = number_of_vacancies_by_year(new_df)
    salary_area = salary_by_area(new_df)
    vacansies_area = number_of_vacancies_by_area(new_df)

    list_tab1 = [[key, round(salary_year[key]), vacansies_year[key]] for key in salary_year]
    list_tab2 = [[key, round(salary_area[key])] for key in salary_area]
    list_tab3 = [[key, vacansies_area[key]] for key in vacansies_area]
    list_tab1.insert(0,["Год","Средняя зарплата","Количество вакансий"])
    list_tab2.insert(0,["Город","Уровень зарплат","Доля вакансий, %"])
    list_tab3.insert(0,["Город","Доля вакансий, %"])

    excel_doc = op.Workbook()
    excel_doc.create_sheet(title='Статистика по годам', index=0)
    excel_doc.create_sheet(title='Статистика по городам', index=1)

    sheetnames = excel_doc.sheetnames
    excel_doc.remove_sheet(excel_doc[sheetnames[2]])

    i=1
    for row in list_tab1:
        if i == 1:
            excel_doc[sheetnames[0]][f"A{i}"] = str(row[0])
            excel_doc[sheetnames[0]][f"B{i}"] = str(row[1])
            excel_doc[sheetnames[0]][f"C{i}"] = str(row[2])
        else:
            excel_doc[sheetnames[0]][f"A{i}"] = row[0]
            excel_doc[sheetnames[0]][f"B{i}"] = row[1]
            excel_doc[sheetnames[0]][f"C{i}"] = row[2]
        i+=1

    i=1
    for row in list_tab2:
        if i == 1:
            excel_doc[sheetnames[1]][f"A{i}"] = str(row[0])
            excel_doc[sheetnames[1]][f"B{i}"] = str(row[1])
        else:
            excel_doc[sheetnames[1]][f"A{i}"] = row[0]
            excel_doc[sheetnames[1]][f"B{i}"] = row[1]
        i+=1
    
    i=1
    for row in list_tab3:
        if i == 1:
            excel_doc[sheetnames[1]][f"D{i}"] = str(row[0])
            excel_doc[sheetnames[1]][f"E{i}"] = str(row[1])
        else:
            excel_doc[sheetnames[1]][f"D{i}"] = row[0]
            excel_doc[sheetnames[1]][f"E{i}"] = row[1]
        i+=1


    excel_doc.save('student_works/report.xlsx')

if __name__ == "__main__":
    create_report()
 